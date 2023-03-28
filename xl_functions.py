#!/usr/bin/python3
# Labor1 - Augusto Burzo

import os
from turtle import color
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
import tkinter.filedialog
from ttkbootstrap.toast import ToastNotification



class XLSaver:
    def __init__(self) -> None:
        self.workbook = Workbook()
        self.worksheet = self.workbook.active

    def customers_list(self, customers) -> None:
        file = [["Labor1 - Lista Clienti"],["Codice cliente", "Ragione sociale","Nome", "Cognome", "Indirizzo", 
                   "Civico", "Provincia", "Citta'", "CAP", "Nazione", "Codice Fiscale",
                   "Partita IVA", "Codice Univoco/Pec", "Telefono", "Cellulare", "FAX",
                   "Email", "Banca", "IBAN", "Listino", "Note"]]
        for customer in customers:
            file.append(customer)
        for row in file:
            self.worksheet.append(row)

        self.worksheet.merge_cells("A1:U1")
        dim_holder = DimensionHolder(worksheet=self.worksheet)

        for col in range(self.worksheet.min_column, self.worksheet.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(self.worksheet, min=col, max=col, width=20)
            self.worksheet.column_dimensions = dim_holder
        try:
            files = [("Excel", "*.xlsx")]
            filename = tkinter.filedialog.asksaveasfilename(filetypes=files, defaultextension=files, title="Lista clienti",
                                                            initialfile="Lista clienti")
            self.workbook.save(filename)
            os.startfile(filename)
            toast = ToastNotification(
                title="Documento salvato",
                message="Il documento è stato salvato con successo!",
                duration=3000,
                icon="☺",
                bootstyle="success"
            )
            toast.show_toast()
        except PermissionError:
            tkinter.messagebox.showerror(title="Impossibile salvare", message="Il file potrebbe essere aperto in un altro"
                                                                         " programma o protetto da scrittura.")

        except FileNotFoundError:
            pass

    def suppliers_list(self, suppliers) -> None:
        file = [["Labor1 - Lista Fornitori"],["Codice cliente", "Ragione sociale", "Indirizzo", 
                   "Civico", "Provincia", "Citta'", "CAP", "Nazione", "Codice Fiscale",
                   "Partita IVA", "Codice Univoco/Pec", "Telefono",
                   "Email", "Banca", "IBAN", "Listino", "Note"]]
        for supplier in suppliers:
            file.append(supplier)
        for row in file:
            self.worksheet.append(row)

        self.worksheet.merge_cells("A1:Q1")
        dim_holder = DimensionHolder(worksheet=self.worksheet)

        for col in range(self.worksheet.min_column, self.worksheet.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(self.worksheet, min=col, max=col, width=20)
            self.worksheet.column_dimensions = dim_holder
        try:
            files = [("Excel", "*.xlsx")]
            filename = tkinter.filedialog.asksaveasfilename(filetypes=files, defaultextension=files, title="Lista clienti",
                                                            initialfile="Lista fornitori")
            self.workbook.save(filename)
            os.startfile(filename)
            toast = ToastNotification(
                title="Documento salvato",
                message="Il documento è stato salvato con successo!",
                duration=3000,
                icon="☺",
                bootstyle="success"
            )
            toast.show_toast()
        except PermissionError:
            tkinter.messagebox.showerror(title="Impossibile salvare", message="Il file potrebbe essere aperto in un altro"
                                                                         " programma o protetto da scrittura.")

        except FileNotFoundError:
            pass

    def agenda(self, events) -> None:
        file = [["Labor1 - Agenda & Scadenze"],["Codice", "Data", "Evento", 
                   "Descrizione", "Contatto", "Stato"]]
        for event in events:
            file.append(event)
        for row in file:
            self.worksheet.append(row)

        self.worksheet.merge_cells("A1:F1")
        dim_holder = DimensionHolder(worksheet=self.worksheet)

        for col in range(2, self.worksheet.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(self.worksheet, min=col, max=col, width=20)
            self.worksheet.column_dimensions = dim_holder
        try:
            files = [("Excel", "*.xlsx")]
            filename = tkinter.filedialog.asksaveasfilename(filetypes=files, defaultextension=files, title="Lista clienti",
                                                            initialfile="Agenda")
            self.workbook.save(filename)
            os.startfile(filename)
            toast = ToastNotification(
                title="Documento salvato",
                message="Il documento è stato salvato con successo!",
                duration=3000,
                icon="☺",
                bootstyle="success"
            )
            toast.show_toast()
        except PermissionError:
            tkinter.messagebox.showerror(title="Impossibile salvare", message="Il file potrebbe essere aperto in un altro"
                                                                         " programma o protetto da scrittura.")

        except FileNotFoundError:
            pass
